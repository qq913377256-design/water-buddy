from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from water_buddy.config import APP_DIR, AppState, total_today_ml


EXCEL_PATH = APP_DIR / "water_log.xlsx"
SUMMARY_SHEET = "每日汇总"
DETAIL_SHEET = "喝水明细"
SUMMARY_HEADERS = ("日期", "目标ml", "实际ml", "完成率", "杯数", "更新时间")
DETAIL_HEADERS = ("日期", "时间", "容量ml")


def sync_water_log(state: AppState, day: str, path: Path = EXCEL_PATH) -> bool:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = load_workbook(path) if path.exists() else Workbook()
        summary = ensure_sheet(workbook, SUMMARY_SHEET, SUMMARY_HEADERS)
        details = ensure_sheet(workbook, DETAIL_SHEET, DETAIL_HEADERS)
        write_summary(summary, state, day)
        write_details(details, state, day)
        workbook.save(path)
        return True
    except PermissionError:
        return False
    except OSError:
        return False


def ensure_sheet(workbook, title: str, headers: tuple[str, ...]):
    if title in workbook.sheetnames:
        sheet = workbook[title]
    else:
        if workbook.active.title == "Sheet" and workbook.active.max_row == 1 and workbook.active.max_column == 1 and workbook.active["A1"].value is None:
            sheet = workbook.active
            sheet.title = title
        else:
            sheet = workbook.create_sheet(title)

    if [sheet.cell(row=1, column=index).value for index in range(1, len(headers) + 1)] != list(headers):
        for index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=index, value=header)
    style_header(sheet, len(headers))
    return sheet


def write_summary(sheet, state: AppState, day: str) -> None:
    row = find_day_row(sheet, day)
    if row is None:
        row = sheet.max_row + 1

    total = total_today_ml(state)
    goal = state.settings.daily_goal_ml
    sheet.cell(row=row, column=1, value=parse_day(day))
    sheet.cell(row=row, column=2, value=goal)
    sheet.cell(row=row, column=3, value=total)
    sheet.cell(row=row, column=4, value=f'=IF(B{row}=0,0,C{row}/B{row})')
    sheet.cell(row=row, column=5, value=len(state.entries))
    sheet.cell(row=row, column=6, value=datetime.now())

    sheet.cell(row=row, column=1).number_format = "yyyy-mm-dd"
    sheet.cell(row=row, column=2).number_format = "#,##0"
    sheet.cell(row=row, column=3).number_format = "#,##0"
    sheet.cell(row=row, column=4).number_format = "0%"
    sheet.cell(row=row, column=5).number_format = "#,##0"
    sheet.cell(row=row, column=6).number_format = "yyyy-mm-dd hh:mm"
    fit_columns(sheet, (12, 10, 10, 10, 8, 18))


def write_details(sheet, state: AppState, day: str) -> None:
    delete_day_detail_rows(sheet, day)
    for entry in state.entries:
        row = sheet.max_row + 1
        created_at = parse_datetime(entry.created_at)
        sheet.cell(row=row, column=1, value=parse_day(day))
        sheet.cell(row=row, column=2, value=created_at)
        sheet.cell(row=row, column=3, value=entry.amount_ml)
        sheet.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        sheet.cell(row=row, column=2).number_format = "hh:mm:ss"
        sheet.cell(row=row, column=3).number_format = "#,##0"
    fit_columns(sheet, (12, 12, 10))


def find_day_row(sheet, day: str) -> int | None:
    for row in range(2, sheet.max_row + 1):
        if normalize_day(sheet.cell(row=row, column=1).value) == day:
            return row
    return None


def delete_day_detail_rows(sheet, day: str) -> None:
    for row in range(sheet.max_row, 1, -1):
        if normalize_day(sheet.cell(row=row, column=1).value) == day:
            sheet.delete_rows(row)


def style_header(sheet, column_count: int) -> None:
    fill = PatternFill("solid", fgColor="DDF4FF")
    for column in range(1, column_count + 1):
        cell = sheet.cell(row=1, column=column)
        cell.font = Font(bold=True, color="244052")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"


def fit_columns(sheet, widths: tuple[int, ...]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def parse_day(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def parse_datetime(value: str) -> datetime:
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return datetime.now()


def normalize_day(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return value[:10]
    return ""
