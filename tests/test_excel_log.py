import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from water_buddy.config import AppState, DrinkEntry, Settings
from water_buddy.excel_log import sync_water_log


class ExcelLogTests(unittest.TestCase):
    def test_sync_writes_summary_and_replaces_day_details(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "water_log.xlsx"
            state = AppState(
                settings=Settings(daily_goal_ml=1600),
                entries=[
                    DrinkEntry(amount_ml=250, created_at="2026-07-07T09:00:00"),
                    DrinkEntry(amount_ml=150, created_at="2026-07-07T10:00:00"),
                ],
            )

            self.assertTrue(sync_water_log(state, "2026-07-07", output))
            state.entries.append(DrinkEntry(amount_ml=300, created_at="2026-07-07T11:00:00"))
            self.assertTrue(sync_water_log(state, "2026-07-07", output))

            workbook = load_workbook(output, data_only=False)
            summary = workbook["每日汇总"]
            details = workbook["喝水明细"]

            self.assertEqual(summary["B2"].value, 1600)
            self.assertEqual(summary["C2"].value, 700)
            self.assertEqual(summary["D2"].value, "=IF(B2=0,0,C2/B2)")
            self.assertEqual(summary["E2"].value, 3)
            self.assertEqual(details.max_row, 4)
            self.assertEqual([details.cell(row=row, column=3).value for row in range(2, 5)], [250, 150, 300])


if __name__ == "__main__":
    unittest.main()
